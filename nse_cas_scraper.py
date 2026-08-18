import json
import logging
import time
from dataclasses import dataclass
from datetime import datetime, time as dt_time
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from curl_cffi import requests


# ============================================================
# CONFIGURATION
# ============================================================

# ------------------------------------------------------------
# CHANGE:
#
# Previous code:
#
# BASE_URL = "https://www.nseindia.com"
#
# New code:
#
# BASE_URL = "https://www.nseindia.com"
#
# No URL change.
#
# The important path-related change is below.
# ------------------------------------------------------------

BASE_URL = "https://www.nseindia.com"

CAS_PAGE_URL = (
    "https://www.nseindia.com/market-data/closing-auction-session"
)

API_ENDPOINT = (
    "https://www.nseindia.com/api/NextApi/apiClient/"
    "casApi?functionName=getCASData"
)


# ============================================================
# NSE FIELD MAPPING
# ============================================================

FIELD_MAP = {
    "symbol": [
        "symbol",
        "SYMBOL",
    ],

    "iep": [
        "IEP",
        "iep",
    ],

    "change": [
        "change",
        "chng",
        "CHNG",
    ],

    "percentage_change": [
        "perChange",
        "percentageChange",
        "pChng",
        "PCHNG",
    ],

    "final_price": [
        "finalPrice",
        "FINALPRICE",
        "final_price",
        "finalprice",
    ],
}


# ============================================================
# RESPONSE STRUCTURE
# ============================================================

ROWS_KEY = "data"


# ============================================================
# TIME CONFIGURATION
# ============================================================

TIMEZONE = ZoneInfo("Asia/Kolkata")

CAS_START_TIME = dt_time(15, 15, 0)
CAS_END_TIME = dt_time(15, 30, 0)

POLL_INTERVAL_SECONDS = 5

REQUEST_TIMEOUT_SECONDS = 10

MAX_RETRIES = 3

RETRY_BACKOFF_SECONDS = 1.0


# ============================================================
# OUTPUT PATHS
# ============================================================

# ------------------------------------------------------------
# CHANGE:
#
# Previous code:
#
# OUTPUT_DIR = Path("output")
#
# New code:
#
# SCRIPT_DIR = Path(__file__).resolve().parent
# OUTPUT_DIR = SCRIPT_DIR / "output"
#
# Reason:
#
# Previously, "output" depended on the directory from which
# Python was started.
#
# For example, if you ran:
#
# python nse_cas_scraper.py
#
# from the project root, it worked.
#
# But if you started the script from another directory,
# Python could create "output" somewhere else.
#
# Now output ALWAYS belongs to:
#
# nse-cas-scraper/
#     output/
#
# regardless of the current PowerShell directory.
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent

OUTPUT_DIR = SCRIPT_DIR / "output"

LOG_DIR = OUTPUT_DIR / "logs"

RAW_DIR = OUTPUT_DIR / "raw"

DATA_DIR = OUTPUT_DIR / "data"


OUTPUT_DIR.mkdir(
    parents=True,
    exist_ok=True,
)

LOG_DIR.mkdir(
    parents=True,
    exist_ok=True,
)

RAW_DIR.mkdir(
    parents=True,
    exist_ok=True,
)

DATA_DIR.mkdir(
    parents=True,
    exist_ok=True,
)


# ============================================================
# OUTPUT FILES
# ============================================================

LATEST_RAW_FILE = (
    RAW_DIR / "cas_raw_latest.json"
)

LATEST_DATA_FILE = (
    DATA_DIR / "cas_latest.json"
)

HISTORY_FILE = (
    DATA_DIR / "cas_history.jsonl"
)

EXCEL_FILE = (
    DATA_DIR / "cas_history.xlsx"
)

EXCEL_SHEET_NAME = "CAS History"

LOG_FILE = (
    LOG_DIR / "nse_cas_scraper.log"
)


# ============================================================
# LOGGING
# ============================================================

logger = logging.getLogger(
    "nse-cas"
)

logger.setLevel(
    logging.INFO
)

logger.handlers.clear()

formatter = logging.Formatter(
    "%(asctime)s | %(levelname)s | %(message)s"
)


console_handler = logging.StreamHandler()

console_handler.setFormatter(
    formatter
)


file_handler = logging.FileHandler(
    LOG_FILE,
    encoding="utf-8",
)

file_handler.setFormatter(
    formatter
)


logger.addHandler(
    console_handler
)

logger.addHandler(
    file_handler
)


# ============================================================
# DATA MODEL
# ============================================================

@dataclass
class CASRecord:

    symbol: str | None

    iep: float | None

    change: float | None

    percentage_change: float | None

    final_price: float | None

    last_update_time: str | None = None

    def as_dict(
        self,
    ) -> dict[str, Any]:

        return {
            "symbol": self.symbol,

            "iep": self.iep,

            "change": self.change,

            "percentage_change": (
                self.percentage_change
            ),

            "final_price": self.final_price,

            "last_update_time": (
                self.last_update_time
            ),
        }


# ============================================================
# SESSION EXCEPTION
# ============================================================

class SessionExpired(Exception):
    pass


# ============================================================
# TIME HELPERS
# ============================================================

def now_ist() -> datetime:

    return datetime.now(
        TIMEZONE
    )


def timestamp_for_filename() -> str:

    return now_ist().strftime(
        "%Y%m%d_%H%M%S_%f"
    )


def collection_window_is_open() -> bool:

    current_time = now_ist().time()

    return (
        CAS_START_TIME
        <= current_time
        < CAS_END_TIME
    )


def wait_until_cas_start() -> None:

    """
    If the program starts before 15:15 IST,
    wait until 15:15 IST.
    """

    now = now_ist()

    if now.time() >= CAS_START_TIME:
        return

    target = datetime.combine(
        now.date(),
        CAS_START_TIME,
        tzinfo=TIMEZONE,
    )

    seconds = (
        target - now
    ).total_seconds()

    logger.info(
        "Current IST time: %s",
        now.strftime("%H:%M:%S"),
    )

    logger.info(
        "CAS collection starts at %s IST.",
        target.strftime("%H:%M:%S"),
    )

    logger.info(
        "Waiting %.1f seconds...",
        seconds,
    )

    time.sleep(
        max(
            0,
            seconds,
        )
    )


# ============================================================
# CREATE NSE SESSION
# ============================================================

def create_session():

    """
    Create a browser-like NSE HTTP session.
    """

    session = requests.Session(
        impersonate="chrome"
    )

    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 "
                "(KHTML, like Gecko) "
                "Chrome/150.0.0.0 Safari/537.36"
            ),

            "Accept": (
                "text/html,application/xhtml+xml,"
                "application/xml;q=0.9,"
                "image/avif,image/webp,"
                "image/apng,*/*;q=0.8"
            ),

            "Accept-Language": (
                "en-IN,en;q=0.9"
            ),

            "Connection": "keep-alive",

            "Cache-Control": "no-cache",

            "Pragma": "no-cache",
        }
    )

    return session


# ============================================================
# WARM UP SESSION
# ============================================================

def warm_up_session(
    session,
) -> None:

    """
    Visit NSE homepage and CAS page before calling API.
    """

    logger.info(
        "Opening NSE homepage..."
    )

    response = session.get(
        BASE_URL,
        timeout=REQUEST_TIMEOUT_SECONDS,
    )

    response.raise_for_status()

    logger.info(
        "NSE homepage loaded. Cookies: %d",
        len(session.cookies),
    )


    logger.info(
        "Opening NSE CAS page..."
    )

    response = session.get(
        CAS_PAGE_URL,
        timeout=REQUEST_TIMEOUT_SECONDS,
        headers={
            "Referer": BASE_URL + "/",
        },
    )

    response.raise_for_status()

    logger.info(
        "NSE CAS page loaded. Cookies: %d",
        len(session.cookies),
    )


# ============================================================
# VALIDATION
# ============================================================

def validate_configuration() -> None:

    if not API_ENDPOINT:

        raise RuntimeError(
            "API_ENDPOINT is empty."
        )

    if not CAS_PAGE_URL:

        raise RuntimeError(
            "CAS_PAGE_URL is empty."
        )

    if ROWS_KEY != "data":

        logger.warning(
            "ROWS_KEY is %r.",
            ROWS_KEY,
        )


# ============================================================
# FETCH API JSON
# ============================================================

def fetch_json(
    session,
) -> dict | list:

    last_exception = None

    for attempt in range(
        1,
        MAX_RETRIES + 1,
    ):

        try:

            logger.info(
                "Fetching CAS API. Attempt %d/%d",
                attempt,
                MAX_RETRIES,
            )

            response = session.get(
                API_ENDPOINT,

                headers={
                    "Accept": (
                        "application/json, "
                        "text/plain, */*"
                    ),

                    "Referer": CAS_PAGE_URL,

                    "Origin": BASE_URL,

                    "User-Agent": session.headers.get(
                        "User-Agent"
                    ),

                    "X-Requested-With": (
                        "XMLHttpRequest"
                    ),
                },

                timeout=REQUEST_TIMEOUT_SECONDS,
            )

            status = response.status_code

            logger.info(
                "CAS API HTTP status: %d",
                status,
            )


            # ------------------------------------------------
            # SESSION EXPIRED
            # ------------------------------------------------

            if status in (
                401,
                403,
            ):

                raise SessionExpired(
                    f"NSE returned HTTP {status}"
                )


            # ------------------------------------------------
            # RATE LIMIT
            # ------------------------------------------------

            if status == 429:

                wait_seconds = (
                    RETRY_BACKOFF_SECONDS
                    * (2 ** (attempt - 1))
                )

                logger.warning(
                    "HTTP 429. Waiting %.1f seconds.",
                    wait_seconds,
                )

                time.sleep(
                    wait_seconds
                )

                continue


            # ------------------------------------------------
            # SERVER ERROR
            # ------------------------------------------------

            if 500 <= status < 600:

                wait_seconds = (
                    RETRY_BACKOFF_SECONDS
                    * (2 ** (attempt - 1))
                )

                logger.warning(
                    "NSE HTTP %d. "
                    "Retrying in %.1f seconds.",
                    status,
                    wait_seconds,
                )

                time.sleep(
                    wait_seconds
                )

                continue


            response.raise_for_status()


            # ------------------------------------------------
            # PARSE JSON
            # ------------------------------------------------

            try:

                raw = response.json()

            except ValueError as exc:

                debug_file = (
                    RAW_DIR
                    / (
                        "invalid_json_"
                        f"{timestamp_for_filename()}.txt"
                    )
                )

                debug_file.write_text(
                    response.text,
                    encoding="utf-8",
                )

                raise RuntimeError(
                    "NSE returned invalid JSON. "
                    f"Response saved to {debug_file}"
                ) from exc


            # ------------------------------------------------
            # RESPONSE DIAGNOSTICS
            # ------------------------------------------------

            if isinstance(
                raw,
                dict,
            ):

                data_rows = raw.get(
                    ROWS_KEY
                )

                symbols = raw.get(
                    "symbols"
                )

                data_count = (
                    len(data_rows)
                    if isinstance(
                        data_rows,
                        list,
                    )
                    else 0
                )

                symbol_count = (
                    len(symbols)
                    if isinstance(
                        symbols,
                        list,
                    )
                    else 0
                )

                logger.info(
                    "CAS API response: "
                    "data_rows=%d, "
                    "available_symbols=%d",
                    data_count,
                    symbol_count,
                )


                if (
                    isinstance(
                        data_rows,
                        list,
                    )
                    and not data_rows
                    and symbol_count > 0
                ):

                    logger.warning(
                        "NSE returned data=[] while "
                        "%d symbols are available. "
                        "No actual CAS records are available "
                        "in this response.",
                        symbol_count,
                    )


                # ------------------------------------------------
                # DEBUG FIRST ROW
                # ------------------------------------------------

                if (
                    isinstance(
                        data_rows,
                        list,
                    )
                    and data_rows
                ):

                    logger.info(
                        "FIRST RAW CAS ROW: %s",

                        json.dumps(
                            data_rows[0],
                            ensure_ascii=False,
                        ),
                    )


            return raw


        except SessionExpired:

            raise


        except (
            requests.exceptions.Timeout,
            requests.exceptions.ConnectionError,
        ) as exc:

            last_exception = exc

            wait_seconds = (
                RETRY_BACKOFF_SECONDS
                * (2 ** (attempt - 1))
            )

            logger.warning(
                "Network error: %s. "
                "Retrying in %.1f seconds.",
                exc,
                wait_seconds,
            )

            time.sleep(
                wait_seconds
            )


    raise RuntimeError(
        "Unable to fetch CAS data after "
        f"{MAX_RETRIES} attempts."
    ) from last_exception


# ============================================================
# ATOMIC JSON WRITE
# ============================================================

def write_json_file(
    path: Path,
    payload: Any,
) -> None:

    """
    Write JSON safely.

    Temporary file is written first and then replaced.

    This prevents the frontend from reading a half-written
    JSON file while the Python scraper is updating it.
    """

    temporary_path = path.with_suffix(
        path.suffix + ".tmp"
    )

    temporary_path.write_text(
        json.dumps(
            payload,
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    temporary_path.replace(
        path
    )


# ============================================================
# SAVE RAW RESPONSE
# ============================================================

def save_raw_response(
    raw: dict | list,
) -> None:

    payload = {
        "timestamp": now_ist().isoformat(),

        "source": API_ENDPOINT,

        "response": raw,
    }


    # --------------------------------------------------------
    # CHANGE:
    #
    # Previous code:
    #
    # LATEST_RAW_FILE.write_text(...)
    #
    # New code:
    #
    # write_json_file(
    #     LATEST_RAW_FILE,
    #     payload,
    # )
    #
    # Reason:
    # Atomic write prevents the frontend from reading an
    # incomplete JSON file during an update.
    # --------------------------------------------------------

    write_json_file(
        LATEST_RAW_FILE,
        payload,
    )


    # Historical raw response

    history_file = (
        RAW_DIR
        / (
            "cas_raw_"
            f"{timestamp_for_filename()}.json"
        )
    )

    write_json_file(
        history_file,
        payload,
    )

    logger.info(
        "Raw response saved: %s",
        history_file,
    )


# ============================================================
# JSON HELPERS
# ============================================================

def get_nested_value(
    row: dict,
    field_path: str,
):

    current = row

    for key in field_path.split("."):

        if not isinstance(
            current,
            dict,
        ):

            return None

        current = current.get(
            key
        )

        if current is None:

            return None

    return current


def get_first_value(
    row: dict,
    field_names: list[str],
):

    for field_name in field_names:

        value = get_nested_value(
            row,
            field_name,
        )

        if value is not None:

            return value

    return None


# ============================================================
# GET ROWS
# ============================================================

def get_rows(
    raw: dict | list,
) -> list[dict]:

    if isinstance(
        raw,
        list,
    ):

        return [
            row
            for row in raw
            if isinstance(
                row,
                dict,
            )
        ]


    if not isinstance(
        raw,
        dict,
    ):

        raise ValueError(
            "Unexpected NSE response type: "
            f"{type(raw).__name__}"
        )


    rows = raw.get(
        ROWS_KEY
    )


    if rows is None:

        logger.warning(
            "Response does not contain key %r.",
            ROWS_KEY,
        )

        logger.warning(
            "Available top-level keys: %s",
            list(raw.keys()),
        )

        return []


    if not isinstance(
        rows,
        list,
    ):

        raise ValueError(
            f"NSE '{ROWS_KEY}' should be a list, "
            f"but got {type(rows).__name__}"
        )


    return [
        row
        for row in rows
        if isinstance(
            row,
            dict,
        )
    ]


# ============================================================
# NUMBER PARSING
# ============================================================

def parse_number(
    value,
) -> float | None:

    if value is None:

        return None


    if isinstance(
        value,
        bool,
    ):

        return None


    if isinstance(
        value,
        (int, float),
    ):

        return float(
            value
        )


    value = str(
        value
    ).strip()


    if not value:

        return None


    if value.upper() in {
        "NA",
        "N/A",
        "NULL",
        "NONE",
        "-",
        "--",
    }:

        return None


    value = (
        value
        .replace(",", "")
        .replace("%", "")
        .strip()
    )


    try:

        return float(
            value
        )

    except ValueError:

        logger.warning(
            "Could not parse numeric value: %r",
            value,
        )

        return None


# ============================================================
# PARSE CAS DATA
# ============================================================

def parse_cas_data(
    raw: dict | list,
) -> list[CASRecord]:

    rows = get_rows(
        raw
    )


    if not rows:

        logger.warning(
            "NSE returned zero CAS rows."
        )

        return []


    records: list[CASRecord] = []


    for index, row in enumerate(
        rows
    ):

        # ----------------------------------------------------
        # SYMBOL
        # ----------------------------------------------------

        symbol_value = get_first_value(
            row,
            FIELD_MAP["symbol"],
        )

        symbol = (
            str(
                symbol_value
            ).strip()
            if symbol_value is not None
            else None
        )


        # ----------------------------------------------------
        # IEP
        # ----------------------------------------------------

        iep = parse_number(
            get_first_value(
                row,
                FIELD_MAP["iep"],
            )
        )


        # ----------------------------------------------------
        # CHANGE
        # ----------------------------------------------------

        change = parse_number(
            get_first_value(
                row,
                FIELD_MAP["change"],
            )
        )


        # ----------------------------------------------------
        # PERCENTAGE CHANGE
        # ----------------------------------------------------

        percentage_change = parse_number(
            get_first_value(
                row,
                FIELD_MAP[
                    "percentage_change"
                ],
            )
        )


        # ----------------------------------------------------
        # FINAL PRICE
        # ----------------------------------------------------

        final_price = parse_number(
            get_first_value(
                row,
                FIELD_MAP["final_price"],
            )
        )


        # ----------------------------------------------------
        # NSE UPDATE TIME
        # ----------------------------------------------------

        last_update_time = get_first_value(
            row,
            [
                "lastUpdateTime",
                "last_update_time",
            ],
        )


        if last_update_time is not None:

            last_update_time = str(
                last_update_time
            ).strip()


        # ----------------------------------------------------
        # CHANGE:
        #
        # Previous code:
        #
        # if final_price == 0.0:
        #     final_price = None
        #
        # New code:
        #
        # Treat all NSE zero placeholders consistently.
        #
        # Reason:
        #
        # After CAS closes, NSE can return:
        #
        # IEP        = 0
        # change     = 0
        # perChange  = 0
        # finalPrice = 0
        #
        # These should not be interpreted as actual market
        # values.
        # ----------------------------------------------------

        if iep == 0.0:

            iep = None


        if change == 0.0:

            change = None


        if percentage_change == 0.0:

            percentage_change = None


        if final_price == 0.0:

            final_price = None


        record = CASRecord(
            symbol=symbol,

            iep=iep,

            change=change,

            percentage_change=(
                percentage_change
            ),

            final_price=final_price,

            last_update_time=(
                last_update_time
            ),
        )


        records.append(
            record
        )


        logger.debug(
            "Parsed row %d: %s",
            index,
            record.as_dict(),
        )


    return records


# ============================================================
# FORMAT VALUES
# ============================================================

def format_number(
    value: float | None,
) -> str:

    if value is None:

        return "-"

    return f"{value:.2f}"


def format_percentage(
    value: float | None,
) -> str:

    if value is None:

        return "-"

    return f"{value:.2f}%"


# ============================================================
# PRINT RECORDS
# ============================================================

def print_records(
    records: list[CASRecord],
) -> None:

    print()

    print(
        "=" * 90
    )

    print(
        "NSE CLOSING AUCTION SESSION"
    )

    print(
        "Timestamp           : "
        f"{now_ist().strftime('%Y-%m-%d %H:%M:%S')} IST"
    )

    print(
        f"Records             : {len(records)}"
    )

    print(
        "=" * 90
    )


    if not records:

        print(
            "No CAS records returned by NSE."
        )

        print(
            "=" * 90
        )

        return


    for record in records:

        print(
            f"Symbol              : "
            f"{record.symbol or '-'}"
        )

        print(
            f"IEP                 : "
            f"{format_number(record.iep)}"
        )

        print(
            f"CHNG                : "
            f"{format_number(record.change)}"
        )

        print(
            f"%CHNG               : "
            f"{format_percentage(record.percentage_change)}"
        )

        print(
            f"Final Price         : "
            f"{format_number(record.final_price)}"
        )

        print(
            f"NSE Update Time     : "
            f"{record.last_update_time or '-'}"
        )

        print(
            "-" * 90
        )


# ============================================================
# SAVE PARSED DATA
# ============================================================

def save_parsed_data(
    records: list[CASRecord],
) -> None:

    timestamp = now_ist()


    payload = {
        "timestamp": timestamp.isoformat(),

        "timestamp_ist": timestamp.strftime(
            "%Y-%m-%d %H:%M:%S"
        ),

        "record_count": len(records),

        "data": [
            record.as_dict()
            for record in records
        ],
    }


    # --------------------------------------------------------
    # CHANGE:
    #
    # Previous code:
    #
    # LATEST_DATA_FILE.write_text(...)
    #
    # New code:
    #
    # write_json_file(
    #     LATEST_DATA_FILE,
    #     payload,
    # )
    #
    # Reason:
    # Prevent frontend from reading incomplete JSON.
    # --------------------------------------------------------

    write_json_file(
        LATEST_DATA_FILE,
        payload,
    )


    # --------------------------------------------------------
    # JSONL HISTORY
    # --------------------------------------------------------

    with HISTORY_FILE.open(
        "a",
        encoding="utf-8",
    ) as file:

        file.write(
            json.dumps(
                payload,
                ensure_ascii=False,
            )
        )

        file.write(
            "\n"
        )


# ============================================================
# SAVE EXCEL HISTORY
# ============================================================

def save_excel_data(
    records: list[CASRecord],
) -> None:

    """
    Append current CAS records to Excel.

    Excel location:

        output/data/cas_history.xlsx
    """

    if not records:

        logger.warning(
            "No records available. "
            "Excel will not be updated."
        )

        return


    # --------------------------------------------------------
    # CHANGE:
    #
    # Previous code:
    #
    # from openpyxl import Workbook, load_workbook
    #
    # New code:
    #
    # Same import, but handled inside this function.
    #
    # --------------------------------------------------------

    from openpyxl import Workbook, load_workbook


    timestamp = now_ist().strftime(
        "%Y-%m-%d %H:%M:%S"
    )


    try:

        # ====================================================
        # CREATE EXCEL FILE
        # ====================================================

        if not EXCEL_FILE.exists():

            logger.info(
                "Creating Excel file: %s",
                EXCEL_FILE,
            )

            workbook = Workbook()

            worksheet = workbook.active

            worksheet.title = (
                EXCEL_SHEET_NAME
            )


            worksheet.append(
                [
                    "timestamp",
                    "symbol",
                    "IEP",
                    "CHNG",
                    "%CHNG",
                    "FINAL PRICE",
                    "NSE UPDATE TIME",
                ]
            )


            workbook.save(
                EXCEL_FILE
            )

            workbook.close()


        # ====================================================
        # OPEN EXCEL
        # ====================================================

        workbook = load_workbook(
            EXCEL_FILE
        )


        # ====================================================
        # GET SHEET
        # ====================================================

        if EXCEL_SHEET_NAME in workbook.sheetnames:

            worksheet = workbook[
                EXCEL_SHEET_NAME
            ]

        else:

            # ------------------------------------------------
            # CHANGE:
            #
            # Previous code:
            #
            # worksheet = workbook["CAS History"]
            #
            # New code:
            #
            # Self-healing sheet lookup.
            #
            # Reason:
            # If the workbook was manually modified,
            # the scraper should not crash because the sheet
            # was renamed.
            # ------------------------------------------------

            worksheet = workbook.active

            worksheet.title = (
                EXCEL_SHEET_NAME
            )


            # If workbook is completely empty,
            # add header.

            if worksheet.max_row == 1:

                existing_values = [
                    cell.value
                    for cell in worksheet[1]
                ]

                if not any(
                    value is not None
                    for value in existing_values
                ):

                    worksheet.append(
                        [
                            "timestamp",
                            "symbol",
                            "IEP",
                            "CHNG",
                            "%CHNG",
                            "FINAL PRICE",
                            "NSE UPDATE TIME",
                        ]
                    )


            logger.warning(
                "Sheet %r was missing. "
                "Active sheet renamed to %r.",
                EXCEL_SHEET_NAME,
                EXCEL_SHEET_NAME,
            )


        # ====================================================
        # APPEND DATA
        # ====================================================

        for record in records:

            worksheet.append(
                [
                    timestamp,

                    record.symbol,

                    record.iep,

                    record.change,

                    record.percentage_change,

                    record.final_price,

                    record.last_update_time,
                ]
            )


        # ====================================================
        # FORMAT NUMBERS
        # ====================================================

        for row in worksheet.iter_rows(
            min_row=2,
            min_col=3,
            max_col=6,
        ):

            for cell in row:

                if cell.value is not None:

                    cell.number_format = (
                        "0.00"
                    )


        # ====================================================
        # HEADER
        # ====================================================

        worksheet.freeze_panes = "A2"


        # ====================================================
        # FILTER
        # ====================================================

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


        # ====================================================
        # COLUMN WIDTH
        # ====================================================

        worksheet.column_dimensions[
            "A"
        ].width = 22

        worksheet.column_dimensions[
            "B"
        ].width = 18

        worksheet.column_dimensions[
            "C"
        ].width = 14

        worksheet.column_dimensions[
            "D"
        ].width = 14

        worksheet.column_dimensions[
            "E"
        ].width = 14

        worksheet.column_dimensions[
            "F"
        ].width = 16

        worksheet.column_dimensions[
            "G"
        ].width = 25


        # ====================================================
        # SAVE
        # ====================================================

        workbook.save(
            EXCEL_FILE
        )

        workbook.close()


        logger.info(
            "Excel updated successfully: "
            "%s | %d stocks | %s",
            EXCEL_FILE,
            len(records),
            timestamp,
        )


    except PermissionError as exc:

        # ----------------------------------------------------
        # IMPORTANT:
        #
        # If cas_history.xlsx is open in Microsoft Excel,
        # Windows can lock the file.
        #
        # The scraper cannot write to a locked Excel file.
        # ----------------------------------------------------

        logger.error(
            "EXCEL FILE IS LOCKED."
        )

        logger.error(
            "Close this file in Microsoft Excel: %s",
            EXCEL_FILE,
        )

        logger.error(
            "Excel error: %s",
            exc,
        )


    except Exception as exc:

        logger.exception(
            "Failed to update Excel: %s",
            exc,
        )


# ============================================================
# SNAPSHOT COMPARISON
# ============================================================

def snapshot_key(
    records: list[CASRecord],
):

    values = []


    for record in records:

        values.append(
            (
                record.symbol,

                record.iep,

                record.change,

                record.percentage_change,

                record.final_price,

                record.last_update_time,
            )
        )


    return tuple(
        sorted(
            values,
            key=lambda x: str(x[0]),
        )
    )


# ============================================================
# COLLECT ONCE
# ============================================================

def collect_once(
    session,
) -> list[CASRecord]:

    # ========================================================
    # FETCH
    # ========================================================

    raw = fetch_json(
        session
    )


    # ========================================================
    # SAVE RAW
    # ========================================================

    save_raw_response(
        raw
    )


    # ========================================================
    # RESPONSE INFORMATION
    # ========================================================

    if isinstance(
        raw,
        dict,
    ):

        logger.info(
            "NSE response keys: %s",
            list(
                raw.keys()
            ),
        )


        if "data" in raw:

            data = raw[
                "data"
            ]

            if isinstance(
                data,
                list,
            ):

                logger.info(
                    "NSE data row count: %d",
                    len(data),
                )


        if "symbols" in raw:

            symbols = raw[
                "symbols"
            ]

            if isinstance(
                symbols,
                list,
            ):

                logger.info(
                    "NSE symbol list count: %d",
                    len(symbols),
                )


    # ========================================================
    # PARSE
    # ========================================================

    records = parse_cas_data(
        raw
    )


    # ========================================================
    # SAVE JSON
    # ========================================================

    save_parsed_data(
        records
    )


    # ========================================================
    # SAVE EXCEL
    # ========================================================

    save_excel_data(
        records
    )


    logger.info(
        "Parsed CAS records: %d",
        len(records),
    )


    return records


# ============================================================
# REFRESH SESSION
# ============================================================

def refresh_session(
    old_session,
):

    try:

        old_session.close()

    except Exception:

        pass


    logger.info(
        "Creating a fresh NSE session..."
    )


    new_session = create_session()


    warm_up_session(
        new_session
    )


    return new_session


# ============================================================
# MAIN COLLECTOR
# ============================================================

def run_collector():

    validate_configuration()


    # ========================================================
    # STARTUP LOG
    # ========================================================

    logger.info(
        "=================================================="
    )

    logger.info(
        "NSE CAS SCRAPER STARTING"
    )

    logger.info(
        "Script directory: %s",
        SCRIPT_DIR,
    )

    logger.info(
        "API endpoint: %s",
        API_ENDPOINT,
    )

    logger.info(
        "CAS window: %s - %s IST",
        CAS_START_TIME.strftime(
            "%H:%M:%S"
        ),
        CAS_END_TIME.strftime(
            "%H:%M:%S"
        ),
    )

    logger.info(
        "Poll interval: %d seconds",
        POLL_INTERVAL_SECONDS,
    )

    logger.info(
        "Output directory: %s",
        OUTPUT_DIR,
    )

    logger.info(
        "Latest JSON: %s",
        LATEST_DATA_FILE,
    )

    logger.info(
        "Raw JSON: %s",
        LATEST_RAW_FILE,
    )

    logger.info(
        "Excel file: %s",
        EXCEL_FILE,
    )

    logger.info(
        "Log file: %s",
        LOG_FILE,
    )

    logger.info(
        "=================================================="
    )


    # ========================================================
    # WAIT FOR 15:15
    # ========================================================

    wait_until_cas_start()


    # ========================================================
    # CHECK CAS WINDOW
    # ========================================================

    if not collection_window_is_open():

        logger.info(
            "CAS collection window is closed."
        )

        logger.info(
            "Current IST time: %s",
            now_ist().strftime(
                "%H:%M:%S"
            ),
        )

        return


    # ========================================================
    # CREATE SESSION
    # ========================================================

    session = create_session()


    try:

        warm_up_session(
            session
        )

    except Exception:

        try:

            session.close()

        except Exception:

            pass

        raise


    previous_snapshot = None


    # ========================================================
    # POLLING LOOP
    # ========================================================

    while collection_window_is_open():

        try:

            records = collect_once(
                session
            )


            current_snapshot = snapshot_key(
                records
            )


            # ------------------------------------------------
            # PRINT
            # ------------------------------------------------

            print_records(
                records
            )


            # ------------------------------------------------
            # CHANGE DETECTION
            # ------------------------------------------------

            if (
                previous_snapshot is None
                or current_snapshot
                != previous_snapshot
            ):

                logger.info(
                    "CAS data changed."
                )

            else:

                logger.info(
                    "CAS data unchanged."
                )


            previous_snapshot = (
                current_snapshot
            )


        # ====================================================
        # SESSION EXPIRED
        # ====================================================

        except SessionExpired as exc:

            logger.warning(
                "NSE session expired: %s",
                exc,
            )


            try:

                session = refresh_session(
                    session
                )

            except Exception as refresh_error:

                logger.exception(
                    "Could not refresh NSE session: %s",
                    refresh_error,
                )

                time.sleep(
                    RETRY_BACKOFF_SECONDS
                )


        # ====================================================
        # ANY OTHER ERROR
        # ====================================================

        except Exception as exc:

            logger.exception(
                "CAS collection failed: %s",
                exc,
            )


        # ====================================================
        # STOP AT 15:30
        # ====================================================

        now = now_ist()


        if now.time() >= CAS_END_TIME:

            break


        end_datetime = datetime.combine(
            now.date(),
            CAS_END_TIME,
            tzinfo=TIMEZONE,
        )


        remaining_seconds = (
            end_datetime - now
        ).total_seconds()


        sleep_seconds = min(
            POLL_INTERVAL_SECONDS,
            max(
                0,
                remaining_seconds,
            ),
        )


        if sleep_seconds > 0:

            logger.info(
                "Next request in %.1f seconds.",
                sleep_seconds,
            )

            time.sleep(
                sleep_seconds
            )


    # ========================================================
    # CLOSE SESSION
    # ========================================================

    try:

        session.close()

    except Exception:

        pass


    logger.info(
        "=================================================="
    )

    logger.info(
        "CAS collection stopped."
    )

    logger.info(
        "Stopped at: %s IST",
        now_ist().strftime(
            "%Y-%m-%d %H:%M:%S"
        ),
    )

    logger.info(
        "=================================================="
    )


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":

    try:

        run_collector()

    except KeyboardInterrupt:

        logger.info(
            "Collector stopped by user."
        )

    except Exception as exc:

        logger.exception(
            "Fatal error: %s",
            exc,
        )